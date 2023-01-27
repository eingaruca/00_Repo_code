from openpyxl import Workbook
from openpyxl.styles import Font,borders,Side, Border
from openpyxl.chart import ScatterChart, Reference, Series
import time, configparser, os, sys, logging, json, random
from pyzabbix.api import ZabbixAPI
import functions

path = os.path.join(os.path.dirname(__file__), 'parameters.txt')
config = configparser.ConfigParser()
config.read(path)
columns = json.loads(config['AXIS']['columns'])
rows = json.loads(config['AXIS']['rows'])
positionx=int(config['AXIS']['positionx'])
positiony=config['AXIS']['positiony']
projects = json.loads(config['AXIS']['projects'])

book = Workbook()
sheet = book.active
sheet2 = book.create_sheet('hoja2')
sheet3 = book.create_sheet('hoja3')
sheet4 = book.create_sheet('hoja4')

sheet['A1']="Ejemplo"

for i in range(2,15):
    sheet[f'A{i}'] = i

for i in range(2,15):
    sheet[f'B{i}'] = i*10

chart1 = ScatterChart()

chart1.title = "Grafico de dispersión"
chart1.style = 13
chart1.y_axis.title = "eje Y"
chart1.x_axis.title = "Eje X"

xvalue = Reference(sheet,min_col=1, min_row=1, max_col=1, max_row=13)
yvalue = Reference(sheet,min_col=2, min_row=1, max_col=2, max_row=13)

ser = Series(yvalue,xvalue,title="Recta")

chart1.series.append(ser)

sheet.add_chart(chart1, "D3")


sheet2['A1'] = 'SUSCRIBETE'
fecha = time.strftime('%x')
sheet2['A2'] = fecha


sheet3.merge_cells('A1:D1')
sheet3['A1'] = 'prueba de union de celdas'
# sheet3.unmerge_cells('A1:D1')

#######################################################
###sheet4
#######################################################

#Cabecera X
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")

'''
for i in range(len(columns)):
    letter=chr(ord("B")+i+1)
    sheet4[f'{letter}{positionx}']=columns[i]
    sheet4[f'{letter}{positionx}'].border = Border(top=double, left=thin, right=thin, bottom=double)
'''

'''
a,b=functions.Functions.print_header_x(sheet4, columns, positionx, positiony)
print("a es", a)
print("b es", b)
#Cabecera Y
#print(type(positiony))
print("positiony", positiony)
c,d = functions.Functions.print_header_y(sheet4, rows, positionx, positiony)
print("c es", c)
print("d es", d)
'''

inix,iniy = functions.Functions.print_all_headers(sheet4, columns, rows, positionx, positiony)
print ("x es ", inix, " -- y es ",iniy)
#Datos

for y in range(len(rows)):
    for x in range(len(columns)):
        letter=chr(ord(iniy)+x+1)
        sheet4[f'{letter}{inix+y+1}'] = random.randint(1,10)

double = Side(border_style="double", color="ff0000")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
#sheet4["B5"].border = double
#for col in columns:
#    print (col)

print("--->" , columns[0])


for new_sheet in projects:
    sheet_temp = book.create_sheet(new_sheet)
    functions.Functions.print_all_headers(sheet_temp, columns, rows, positionx, positiony)
    for y in range(len(rows)):
        for x in range(len(columns)):
            letter=chr(ord(iniy)+x)
            sheet_temp[f'{letter}{inix+y}'] = random.randint(1,10)
    functions.Functions.create_graph("Gráfico x",sheet_temp, "H", "20")
    


book.save("prueba.xlsx")

