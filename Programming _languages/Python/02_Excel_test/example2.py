from datetime import datetime
import openpyxl
from openpyxl.styles import Font

table_name = ["M-1", "M-2", "M-3"]

xltitle = datetime.now()
tit_head = "ALL_MACHINE" + xltitle.strftime("%d_%m_%Y_%H_%M_%S")+".xlsx"

filepath = tit_head
headings = ("NAME", "ID", "EMPLOYEE NAME",
            "NUMBER", "START TIME", "STOP TIME")
wb = openpyxl.Workbook()
for tab_nam in table_name:
    sheet = wb.create_sheet()
    sheet.title = tab_nam

    sheet.row_dimensions[1].font = Font(bold=True)

    for colno, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=colno).value = heading
wb.save(filepath)